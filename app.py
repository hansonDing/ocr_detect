import os
import json
import zipfile
import shutil
from datetime import datetime
from flask import Flask, request, render_template, jsonify, send_file
from werkzeug.utils import secure_filename
from PIL import Image
import base64
from openai import OpenAI
from dotenv import load_dotenv
from database import init_database, extract_key_information, save_to_database, get_all_records, search_records, delete_record

# PDF processing imports
try:
    import fitz  # PyMuPDF
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("Warning: PyMuPDF not installed. PDF processing will be disabled.")

try:
    from pdf2image import convert_from_path
    PDF2IMAGE_SUPPORT = True
except ImportError:
    PDF2IMAGE_SUPPORT = False
    print("Warning: pdf2image not installed. Alternative PDF processing will be used.")

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Please change to a secure key

# Configuration
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp', 'pdf', 'zip'}
IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB

# Create necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Initialize database
init_database()

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# OpenAI API configuration - for calling ModelScope OCR
OCR_API_BASE = os.getenv('OCR_API_BASE', 'http://localhost:8000/v1')
OCR_API_KEY = os.getenv('OCR_API_KEY', 'EMPTY')
OCR_MODEL = os.getenv('OCR_MODEL', 'OCR')

# Initialize OpenAI client for OCR
print("Initializing OCR client...")
ocr_client = None
ocr_status = "Not initialized"

try:
    ocr_client = OpenAI(
        api_key=OCR_API_KEY,
        base_url=OCR_API_BASE
    )
    
    # Test connection
    try:
        models = ocr_client.models.list()
        print(f"OCR API connection successful! Available models: {[model.id for model in models.data]}")
        ocr_status = "Connected"
    except Exception as e:
        print(f"OCR API connection test failed: {e}")
        ocr_status = "Failed"
        ocr_client = None
except Exception as e:
    print(f"OCR client initialization failed: {e}")
    ocr_status = "Initialization failed"
    ocr_client = None

# Backup OCR initialization
backup_ocr = None
try:
    import pytesseract
    backup_ocr = "tesseract"
    print("Tesseract backup OCR available")
except ImportError:
    print("Tesseract not available")

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_image_file(filename):
    """Check if file is an image"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in IMAGE_EXTENSIONS

def is_pdf_file(filename):
    """Check if file is a PDF"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() == 'pdf'

def convert_pdf_to_images(pdf_path, output_dir):
    """Convert PDF to images using available libraries"""
    try:
        images = []
        
        if PDF2IMAGE_SUPPORT:
            # Method 1: Use pdf2image (requires poppler)
            try:
                pages = convert_from_path(pdf_path, dpi=200)
                for i, page in enumerate(pages):
                    image_filename = f"page_{i+1}.png"
                    image_path = os.path.join(output_dir, image_filename)
                    page.save(image_path, 'PNG')
                    images.append(image_path)
                return images, None
            except Exception as e:
                print(f"pdf2image failed: {e}, trying PyMuPDF...")
        
        if PDF_SUPPORT:
            # Method 2: Use PyMuPDF
            try:
                doc = fitz.open(pdf_path)
                for page_num in range(len(doc)):
                    page = doc.load_page(page_num)
                    # Render page to image
                    mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better quality
                    pix = page.get_pixmap(matrix=mat)
                    image_filename = f"page_{page_num+1}.png"
                    image_path = os.path.join(output_dir, image_filename)
                    pix.save(image_path)
                    images.append(image_path)
                doc.close()
                return images, None
            except Exception as e:
                return [], f"PyMuPDF conversion failed: {str(e)}"
        
        return [], "No PDF processing library available. Please install PyMuPDF or pdf2image."
        
    except Exception as e:
        return [], f"PDF conversion failed: {str(e)}"

def process_pdf_file(pdf_path, filename, output_format='json'):
    """Process PDF file by converting to images and performing OCR"""
    try:
        # Create temporary directory for PDF images
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_pdf_dir = os.path.join(UPLOAD_FOLDER, f"pdf_temp_{timestamp}")
        os.makedirs(temp_pdf_dir, exist_ok=True)
        
        # Convert PDF to images
        image_files, error = convert_pdf_to_images(pdf_path, temp_pdf_dir)
        if error:
            return None, error
        
        if not image_files:
            return None, "No pages found in PDF file"
        
        # Create output directory for this PDF
        pdf_name_without_ext = os.path.splitext(filename)[0]
        output_dir = os.path.join(OUTPUT_FOLDER, f"{pdf_name_without_ext}_{timestamp}")
        os.makedirs(output_dir, exist_ok=True)
        
        # Process each page
        results = []
        combined_text = []
        
        for i, image_path in enumerate(image_files):
            try:
                page_num = i + 1
                page_filename = f"page_{page_num}"
                
                # Perform OCR on this page
                ocr_result, model_used = perform_ocr(image_path)
                combined_text.append(f"=== Page {page_num} ===\n{ocr_result}\n")
                
                # Save individual page result
                original_output_folder = OUTPUT_FOLDER
                globals()['OUTPUT_FOLDER'] = output_dir
                
                try:
                    result_path = save_result_to_file(
                        page_filename, 
                        ocr_result, 
                        f"{filename} - Page {page_num}", 
                        model_used, 
                        output_format
                    )
                    result_filename = os.path.basename(result_path)
                finally:
                    globals()['OUTPUT_FOLDER'] = original_output_folder
                
                results.append({
                    "page_number": page_num,
                    "result_file": result_filename,
                    "output_format": output_format,
                    "ocr_result": ocr_result,
                    "model_used": model_used,
                    "status": "success" if "failed" not in ocr_result else "failed"
                })
                
            except Exception as e:
                results.append({
                    "page_number": i + 1,
                    "result_file": None,
                    "output_format": output_format,
                    "ocr_result": f"Processing failed: {str(e)}",
                    "model_used": "None",
                    "status": "failed"
                })
        
        # Save combined result for all pages
        combined_ocr_text = "\n".join(combined_text)
        original_output_folder = OUTPUT_FOLDER
        globals()['OUTPUT_FOLDER'] = output_dir
        
        try:
            combined_result_path = save_result_to_file(
                f"{pdf_name_without_ext}_combined", 
                combined_ocr_text, 
                f"{filename} - All Pages", 
                results[0]["model_used"] if results else "Unknown", 
                output_format
            )
            combined_result_filename = os.path.basename(combined_result_path)
        finally:
            globals()['OUTPUT_FOLDER'] = original_output_folder
        
        # Clean up temporary directory
        try:
            shutil.rmtree(temp_pdf_dir)
        except:
            pass
        
        return {
            "output_directory": os.path.basename(output_dir),
            "total_pages": len(image_files),
            "processed_results": results,
            "combined_result_file": combined_result_filename,
            "combined_ocr_text": combined_ocr_text,
            "success_count": len([r for r in results if r["status"] == "success"]),
            "failed_count": len([r for r in results if r["status"] == "failed"])
        }, None
        
    except Exception as e:
        # Clean up on error
        try:
            if 'temp_pdf_dir' in locals():
                shutil.rmtree(temp_pdf_dir)
        except:
            pass
        return None, f"PDF processing failed: {str(e)}"

def extract_zip_file(zip_path, extract_to):
    """Extract ZIP file and return list of image files"""
    try:
        image_files = []
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            # Get list of files in ZIP
            file_list = zip_ref.namelist()
            
            # Filter image files
            for file_name in file_list:
                if is_image_file(file_name) and not file_name.startswith('__MACOSX/'):
                    # Extract file
                    zip_ref.extract(file_name, extract_to)
                    extracted_path = os.path.join(extract_to, file_name)
                    image_files.append(extracted_path)
        
        return image_files, None
    except Exception as e:
        return [], f"ZIP extraction failed: {str(e)}"

def process_zip_batch(zip_path, zip_filename, output_format='json'):
    """Process all images in a ZIP file"""
    try:
        # Create temporary extraction directory
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_extract_dir = os.path.join(UPLOAD_FOLDER, f"temp_{timestamp}")
        os.makedirs(temp_extract_dir, exist_ok=True)
        
        # Extract ZIP file
        image_files, error = extract_zip_file(zip_path, temp_extract_dir)
        if error:
            return None, error
        
        if not image_files:
            return None, "No image files found in ZIP archive"
        
        # Create output directory for this ZIP
        zip_name_without_ext = os.path.splitext(zip_filename)[0]
        output_dir = os.path.join(OUTPUT_FOLDER, f"{zip_name_without_ext}_{timestamp}")
        os.makedirs(output_dir, exist_ok=True)
        
        # Process each image
        results = []
        for image_path in image_files:
            try:
                # Get relative filename for JSON naming
                rel_path = os.path.relpath(image_path, temp_extract_dir)
                filename_without_ext = os.path.splitext(os.path.basename(rel_path))[0]
                
                # Perform OCR
                ocr_result, model_used = perform_ocr(image_path)
                
                # Save result to specified format in the ZIP-specific output directory
                # Temporarily change OUTPUT_FOLDER to the ZIP-specific directory
                original_output_folder = OUTPUT_FOLDER
                globals()['OUTPUT_FOLDER'] = output_dir
                
                try:
                    result_path = save_result_to_file(filename_without_ext, ocr_result, rel_path, model_used, output_format)
                    result_filename = os.path.basename(result_path)
                finally:
                    # Restore original OUTPUT_FOLDER
                    globals()['OUTPUT_FOLDER'] = original_output_folder
                
                results.append({
                    "image_file": rel_path,
                    "result_file": result_filename,
                    "output_format": output_format,
                    "ocr_result": ocr_result,
                    "model_used": model_used,
                    "status": "success" if "failed" not in ocr_result else "failed"
                })
                
            except Exception as e:
                results.append({
                    "image_file": os.path.relpath(image_path, temp_extract_dir),
                    "result_file": None,
                    "output_format": output_format,
                    "ocr_result": f"Processing failed: {str(e)}",
                    "model_used": "None",
                    "status": "failed"
                })
        
        # Clean up temporary directory
        try:
            shutil.rmtree(temp_extract_dir)
        except:
            pass
        
        return {
            "output_directory": os.path.basename(output_dir),
            "total_images": len(image_files),
            "processed_results": results,
            "success_count": len([r for r in results if r["status"] == "success"]),
            "failed_count": len([r for r in results if r["status"] == "failed"])
        }, None
        
    except Exception as e:
        # Clean up on error
        try:
            if 'temp_extract_dir' in locals():
                shutil.rmtree(temp_extract_dir)
        except:
            pass
        return None, f"ZIP processing failed: {str(e)}"

def image_to_base64(image_path):
    """Convert image to base64 encoding"""
    try:
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except Exception as e:
        raise Exception(f"Image encoding failed: {str(e)}")

def ocr_with_api(image_path, image_type='document'):
    """Call OCR model using OpenAI API format"""
    try:
        if not ocr_client:
            return "OCR client not initialized"
        
        base64_image = image_to_base64(image_path)
        
        # Choose prompt based on image type
        if image_type == 'table':
            prompt_text = "Please recognize all content in this table image and output it in a structured table format. Requirements:\n1. Identify all table headers, rows, and columns accurately\n2. Preserve the exact table structure and cell relationships\n3. Use consistent spacing or tab characters to align columns\n4. Maintain the original row order and column sequence\n5. Include all cell content including numbers, text, and symbols\n6. For merged cells, indicate the span appropriately\n7. Preserve any table formatting like borders or separators using text characters (|, -, +)\n8. Output should be readable as a plain text table that maintains the original structure\n9. Include table headers if present\n10. Ensure the output can be easily converted to CSV or Excel format"
        else:  # document type
            prompt_text = "Please recognize all text content in this image and output it exactly as it appears in the original image layout. Requirements:\n1. Preserve the exact spatial layout and formatting of the original image\n2. Maintain all spacing, indentation, line breaks, and alignment as shown\n3. For tables: preserve column alignment and row structure using spaces or tabs\n4. For multi-column text: maintain the column layout and reading order\n5. For titles, headings, and special formatting: preserve their visual hierarchy and positioning\n6. Keep all original punctuation, symbols, and special characters\n7. Do not add any explanatory text or formatting markers - output only the recognized content in its original layout\n8. If text appears in different sizes or styles, maintain the relative positioning but output as plain text\n9. Preserve any mathematical formulas or equations in their original format\n10. Ensure the output can be directly used to recreate the visual layout of the original document"
        
        response = ocr_client.chat.completions.create(
            model=OCR_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": prompt_text
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{base64_image}"
                            }
                        }
                    ]
                }
            ],
            max_tokens=2000,
            temperature=0.1
        )
        
        if response.choices and len(response.choices) > 0:
            content = response.choices[0].message.content
            if content:
                return content.strip()
            else:
                return "OCR returned no recognition result"
        else:
            return "OCR API response format error"
            
    except Exception as e:
        return f"OCR API call failed: {str(e)}"

def ocr_with_tesseract(image_path):
    """Use Tesseract for OCR recognition"""
    try:
        import pytesseract
        
        image = Image.open(image_path)
        text = pytesseract.image_to_string(image, lang='chi_sim+eng')
        
        if text.strip():
            return text.strip()
        else:
            return "No text content recognized"
            
    except Exception as e:
        return f"Tesseract OCR recognition failed: {str(e)}"

def ocr_with_basic(image_path):
    """Basic OCR solution (returns image information only)"""
    try:
        with Image.open(image_path) as img:
            width, height = img.size
            mode = img.mode
            format_name = img.format
            
        return f"Image processed successfully\nFile path: {image_path}\nImage dimensions: {width}x{height}\nColor mode: {mode}\nFile format: {format_name}\n\nNote: Currently using basic image processing mode, no text recognition performed.\nRecommend configuring OCR API or installing Tesseract to enable text recognition."
        
    except Exception as e:
        return f"Basic image processing failed: {str(e)}"

def perform_ocr(image_path, image_type='document'):
    """Perform OCR recognition, try different solutions by priority"""
    # Solution 1: OCR API
    if ocr_client and ocr_status == "Connected":
        result = ocr_with_api(image_path, image_type)
        if "failed" not in result and "error" not in result:
            return result, "OCR API"
    
    # Solution 2: Tesseract
    if backup_ocr == "tesseract":
        result = ocr_with_tesseract(image_path)
        if "failed" not in result:
            return result, "Tesseract"
    
    # Solution 3: Basic processing
    result = ocr_with_basic(image_path)
    return result, "Basic"

def process_ocr_and_extract_info(image_path, filename, image_type='document'):
    """Perform OCR recognition and extract key information to database"""
    try:
        # Perform OCR recognition
        ocr_result, model_used = perform_ocr(image_path, image_type)
        
        # Extract key information
        extracted_info = extract_key_information(ocr_result)
        
        # Save to database
        record_id = save_to_database(
            filename=filename,
            file_path=image_path,
            ocr_text=ocr_result,
            model_used=model_used,
            extracted_info=extracted_info
        )
        
        return {
            'ocr_result': ocr_result,
            'model_used': model_used,
            'extracted_info': extracted_info,
            'record_id': record_id,
            'success': True
        }
        
    except Exception as e:
        return {
            'ocr_result': f"Processing failed: {str(e)}",
            'model_used': "None",
            'extracted_info': {},
            'record_id': None,
            'success': False,
            'error': str(e)
        }

def save_result_to_file(filename, ocr_result, original_filename, model_used, output_format='json'):
    """Save OCR result to specified format file"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if output_format == 'json':
        result_filename = f"{filename}_{timestamp}.json"
        result_path = os.path.join(OUTPUT_FOLDER, result_filename)
        
        result_data = {
            "timestamp": datetime.now().isoformat(),
            "original_filename": original_filename,
            "ocr_result": ocr_result,
            "processing_info": {
                "model": model_used,
                "status": "success" if "failed" not in ocr_result else "failed",
                "api_endpoint": OCR_API_BASE if model_used == "OCR API" else "Local",
                "model_name": OCR_MODEL if model_used == "OCR API" else model_used
            }
        }
        
        with open(result_path, 'w', encoding='utf-8') as f:
            json.dump(result_data, f, ensure_ascii=False, indent=2)
            
    elif output_format == 'txt':
        result_filename = f"{filename}_{timestamp}.txt"
        result_path = os.path.join(OUTPUT_FOLDER, result_filename)
        
        with open(result_path, 'w', encoding='utf-8') as f:
            f.write(f"Recognition Time: {datetime.now().isoformat()}\n")
            f.write(f"Original Filename: {original_filename}\n")
            f.write(f"Recognition Model: {model_used}\n")
            f.write(f"Processing Status: {'Success' if 'failed' not in ocr_result else 'Failed'}\n")
            f.write("\n")
            f.write("Recognition Result:\n")
            f.write(ocr_result)
            
    elif output_format == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            
            result_filename = f"{filename}_{timestamp}.xlsx"
            result_path = os.path.join(OUTPUT_FOLDER, result_filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "OCR Recognition Result"
            
            # Set title style
            title_font = Font(bold=True, size=12)
            
            # Add headers and data
            ws['A1'] = "Item"
            ws['B1'] = "Content"
            ws['A1'].font = title_font
            ws['B1'].font = title_font
            
            ws['A2'] = "Recognition Time"
            ws['B2'] = datetime.now().isoformat()
            
            ws['A3'] = "Original Filename"
            ws['B3'] = original_filename
            
            ws['A4'] = "Recognition Model"
            ws['B4'] = model_used
            
            ws['A5'] = "Processing Status"
            ws['B5'] = "Success" if "failed" not in ocr_result else "Failed"
            
            ws['A6'] = "Recognition Result"
            ws['B6'] = ocr_result
            
            # Set column width
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 50
            
            # Set text wrapping
            for row in ws.iter_rows(min_row=2, max_row=6, min_col=2, max_col=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            wb.save(result_path)
            
        except ImportError:
            # If openpyxl is not installed, fallback to JSON format
            return save_result_to_file(filename, ocr_result, original_filename, model_used, 'json')
    
    return result_path

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and OCR recognition"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file selected'})
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename_without_ext = os.path.splitext(filename)[0]
            file_extension = os.path.splitext(filename)[1].lower()
            new_filename = f"{filename_without_ext}_{timestamp}{file_extension}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], new_filename)
            file.save(file_path)
            
            # Check if uploaded file is a ZIP
            if file_extension == '.zip':
                # Get output format from form data
                output_format = request.form.get('output_format', 'json')
                
                # Process ZIP file
                batch_result, error = process_zip_batch(file_path, filename, output_format)
                
                if error:
                    return jsonify({
                        'success': False,
                        'message': f'ZIP processing failed: {error}'
                    })
                
                return jsonify({
                    'success': True,
                    'message': f'ZIP file processed successfully in {output_format.upper()} format. {batch_result["success_count"]} images processed, {batch_result["failed_count"]} failed.',
                    'file_type': 'zip',
                    'output_format': output_format,
                    'uploaded_file': new_filename,
                    'output_directory': batch_result['output_directory'],
                    'batch_results': {
                        'total_images': batch_result['total_images'],
                        'success_count': batch_result['success_count'],
                        'failed_count': batch_result['failed_count'],
                        'processed_results': batch_result['processed_results']
                    }
                })
            # Check if uploaded file is a PDF
            elif file_extension == '.pdf':
                # Check if PDF processing is available
                if not PDF_SUPPORT and not PDF2IMAGE_SUPPORT:
                    return jsonify({
                        'success': False,
                        'message': 'PDF processing not available. Please install PyMuPDF or pdf2image library.'
                    })
                
                # Get output format from form data
                output_format = request.form.get('output_format', 'json')
                
                # Process PDF file
                pdf_result, error = process_pdf_file(file_path, filename, output_format)
                
                if error:
                    return jsonify({
                        'success': False,
                        'message': f'PDF processing failed: {error}'
                    })
                
                # Extract key information from combined OCR text and save to database
                extracted_info = None
                record_id = None
                extracted_count = 0
                
                if pdf_result and pdf_result.get('combined_ocr_text'):
                    try:
                        # Extract key information from combined OCR text
                        extracted_info = extract_key_information(pdf_result['combined_ocr_text'])
                        
                        # Save to database
                        record_id = save_to_database(
                            filename=filename,
                            file_path=file_path,
                            ocr_text=pdf_result['combined_ocr_text'],
                            extracted_info=extracted_info,
                            model_used=pdf_result['processed_results'][0]['model_used'] if pdf_result['processed_results'] else 'Unknown'
                        )
                        
                        # Count extracted fields
                        extracted_count = sum(1 for v in extracted_info.values() if v is not None)
                        
                    except Exception as e:
                        print(f"Database storage failed for PDF: {str(e)}")
                
                return jsonify({
                    'success': True,
                    'message': f'PDF file processed successfully in {output_format.upper()} format. {pdf_result["success_count"]} pages processed, {pdf_result["failed_count"]} failed. {extracted_count} key fields extracted and saved to database.',
                    'file_type': 'pdf',
                    'output_format': output_format,
                    'uploaded_file': new_filename,
                    'output_directory': pdf_result['output_directory'],
                    'combined_result_file': pdf_result['combined_result_file'],
                    'extracted_info': extracted_info,
                    'record_id': record_id,
                    'extracted_fields_count': extracted_count,
                    'pdf_results': {
                        'total_pages': pdf_result['total_pages'],
                        'success_count': pdf_result['success_count'],
                        'failed_count': pdf_result['failed_count'],
                        'processed_results': pdf_result['processed_results'],
                        'combined_ocr_text': pdf_result['combined_ocr_text'][:500] + '...' if len(pdf_result['combined_ocr_text']) > 500 else pdf_result['combined_ocr_text']
                    }
                })
            else:
                # Get output format and image type from form data
                output_format = request.form.get('output_format', 'json')
                image_type = request.form.get('image_type', 'document')
                
                # Process single image file with key information extraction
                processing_result = process_ocr_and_extract_info(file_path, filename, image_type)
                
                if processing_result['success']:
                    # Save result to file
                    result_path = save_result_to_file(
                        filename_without_ext, 
                        processing_result['ocr_result'], 
                        filename, 
                        processing_result['model_used'], 
                        output_format
                    )
                    
                    # Count extracted fields
                    extracted_count = sum(1 for v in processing_result['extracted_info'].values() if v is not None)
                    
                    return jsonify({
                        'success': True,
                        'message': f'OCR recognition completed, saved as {output_format.upper()} format. {extracted_count} key fields extracted and saved to database.',
                        'file_type': 'image',
                        'ocr_result': processing_result['ocr_result'],
                        'extracted_info': processing_result['extracted_info'],
                        'record_id': processing_result['record_id'],
                        'extracted_fields_count': extracted_count,
                        'result_file': os.path.basename(result_path),
                        'output_format': output_format,
                        'uploaded_file': new_filename,
                        'model_info': processing_result['model_used'],
                        'api_info': {
                            'endpoint': OCR_API_BASE if processing_result['model_used'] == "OCR API" else "Local",
                            'model': OCR_MODEL if processing_result['model_used'] == "OCR API" else processing_result['model_used']
                        }
                    })
                else:
                    return jsonify({
                        'success': False,
                        'message': f'OCR processing failed: {processing_result.get("error", "Unknown error")}',
                        'file_type': 'image',
                        'uploaded_file': new_filename
                    })
            
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Processing failed: {str(e)}'
            })
    
    else:
        return jsonify({
            'success': False,
            'message': 'Unsupported file format. Please upload image files (PNG, JPG, JPEG, GIF, BMP, WEBP) or PDF files'
        })

@app.route('/results')
def list_results():
    """List all OCR result files and batch processing directories"""
    try:
        results = {
            'single_files': [],
            'batch_directories': []
        }
        
        # Scan output directory
        for item in os.listdir(OUTPUT_FOLDER):
            item_path = os.path.join(OUTPUT_FOLDER, item)
            
            if os.path.isfile(item_path) and (item.endswith('.json') or item.endswith('.txt') or item.endswith('.xlsx')):
                # Single result files (JSON, TXT, XLSX)
                results['single_files'].append(item)
            elif os.path.isdir(item_path):
                # Batch processing directories
                dir_info = {
                    'name': item,
                    'files': [],
                    'created_time': os.path.getctime(item_path)
                }
                
                # List result files in the directory (JSON, TXT, XLSX)
                try:
                    for file in os.listdir(item_path):
                        if file.endswith('.json') or file.endswith('.txt') or file.endswith('.xlsx'):
                            file_path = os.path.join(item_path, file)
                            dir_info['files'].append({
                                'name': file,
                                'size': os.path.getsize(file_path),
                                'modified_time': os.path.getmtime(file_path)
                            })
                    
                    # Sort files by modification time (newest first)
                    dir_info['files'].sort(key=lambda x: x['modified_time'], reverse=True)
                    results['batch_directories'].append(dir_info)
                except Exception as e:
                    print(f"Error reading directory {item}: {e}")
        
        # Sort single files and directories by creation time (newest first)
        results['single_files'].sort(reverse=True)
        results['batch_directories'].sort(key=lambda x: x['created_time'], reverse=True)
        
        return jsonify({
            'success': True,
            'results': results
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to get result list: {str(e)}'
        })

@app.route('/download/<filename>')
def download_file(filename):
    """Download result file from root output directory"""
    try:
        # Security check: ensure filename contains only supported file formats
        if not (filename.endswith('.json') or filename.endswith('.txt') or filename.endswith('.xlsx')):
            return jsonify({'success': False, 'message': 'Invalid file type'}), 400
        
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        
        # Check if file exists
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'File not found'}), 404
        
        # Set MIME type based on file type
        if filename.endswith('.json'):
            mimetype = 'application/json'
        elif filename.endswith('.txt'):
            mimetype = 'text/plain'
        elif filename.endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Send file
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Download failed: {str(e)}'
        }), 500

@app.route('/download/<directory>/<filename>')
def download_batch_file(directory, filename):
    """Download result file from batch processing directory"""
    try:
        # Security check: ensure filename contains only supported file formats
        if not (filename.endswith('.json') or filename.endswith('.txt') or filename.endswith('.xlsx')):
            return jsonify({'success': False, 'message': 'Invalid file type'}), 400
        
        # Security check: prevent path traversal attacks
        if '..' in directory or '/' in directory or '\\' in directory:
            return jsonify({'success': False, 'message': 'Invalid directory name'}), 400
        
        # Build file path
        dir_path = os.path.join(OUTPUT_FOLDER, directory)
        file_path = os.path.join(dir_path, filename)
        
        # Check if directory exists
        if not os.path.exists(dir_path) or not os.path.isdir(dir_path):
            return jsonify({'success': False, 'message': 'Directory not found'}), 404
        
        # Check if file exists
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'File not found'}), 404
        
        # Set MIME type based on file type
        if filename.endswith('.json'):
            mimetype = 'application/json'
        elif filename.endswith('.txt'):
            mimetype = 'text/plain'
        elif filename.endswith('.xlsx'):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Send file
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Download failed: {str(e)}'
        }), 500

@app.route('/model-status')
def model_status():
    """Check model status"""
    return jsonify({
        'ocr_api': {
            'status': ocr_status,
            'api_base': OCR_API_BASE,
            'model': OCR_MODEL,
            'available': ocr_client is not None and ocr_status == "Connected"
        },
        'backup_ocr': {
            'tesseract': backup_ocr == "tesseract",
            'basic': True
        },
        'current_priority': [
            "OCR API" if ocr_status == "Connected" else None,
            "Tesseract" if backup_ocr == "tesseract" else None,
            "Basic"
        ]
    })

@app.route('/database/records')
def get_database_records():
    """Get all database records"""
    try:
        records = get_all_records()
        return jsonify({
            'success': True,
            'records': records,
            'total_count': len(records)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to get database records: {str(e)}'
        })

@app.route('/database/search')
def search_database_records():
    """Search database records"""
    try:
        customer_name = request.args.get('customer_name', '')
        customer_id = request.args.get('customer_id', '')
        transaction_id = request.args.get('transaction_id', '')
        customer_country = request.args.get('customer_country', '')
        
        records = search_records(
            customer_name=customer_name if customer_name else None,
            customer_id=customer_id if customer_id else None,
            transaction_id=transaction_id if transaction_id else None,
            customer_country=customer_country if customer_country else None
        )
        
        return jsonify({
            'success': True,
            'records': records,
            'total_count': len(records),
            'search_params': {
                'customer_name': customer_name,
                'customer_id': customer_id,
                'transaction_id': transaction_id,
                'customer_country': customer_country
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Search failed: {str(e)}'
        })

@app.route('/database/record/<int:record_id>', methods=['DELETE'])
def delete_database_record(record_id):
    """Delete specified database record"""
    try:
        success = delete_record(record_id)
        if success:
            return jsonify({
                'success': True,
                'message': f'Record {record_id} deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': f'Record {record_id} not found'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Delete failed: {str(e)}'
        }), 500

@app.route('/api/database/stats')
def get_database_stats():
    """Get database statistics"""
    try:
        from database import get_database_stats
        stats = get_database_stats()
        return jsonify(stats)
    except Exception as e:
        return jsonify({
            'total': 0,
            'extracted': 0,
            'today': 0,
            'error': str(e)
        })

@app.route('/api/database/search')
def advanced_search_records():
    """Advanced search database records"""
    try:
        from database import advanced_search_records
        
        # Get search parameters
        keyword = request.args.get('keyword', '')
        customer_name = request.args.get('customer_name', '')
        min_amount = request.args.get('min_amount')
        max_amount = request.args.get('max_amount')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Convert numeric parameters
        min_amount = float(min_amount) if min_amount else None
        max_amount = float(max_amount) if max_amount else None
        
        records = advanced_search_records(
            keyword=keyword if keyword else None,
            customer_name=customer_name if customer_name else None,
            min_amount=min_amount,
            max_amount=max_amount,
            start_date=start_date if start_date else None,
            end_date=end_date if end_date else None
        )
        
        return jsonify({
            'success': True,
            'records': records,
            'total_count': len(records)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Advanced search failed: {str(e)}',
            'records': []
        })

@app.route('/api/database/export')
def export_database_records():
    """Export database records as CSV file"""
    try:
        import csv
        import io
        from flask import make_response
        
        # Get all records
        records = get_all_records()
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        headers = [
            'ID', 'Filename', 'Processing Time', 'Customer Name', 'Transaction Amount', 'Transaction Date', 
            'Transaction ID', 'Account Number', 'Bank Name', 'OCR Model', 'Confidence', 'OCR Raw Text'
        ]
        writer.writerow(headers)
        
        # Write data
        for record in records:
            row = [
                record.get('id', ''),
                record.get('filename', ''),
                record.get('processing_time', ''),
                record.get('customer_name', ''),
                record.get('transaction_amount', ''),
                record.get('transaction_date', ''),
                record.get('transaction_id', ''),
                record.get('account_number', ''),
                record.get('bank_name', ''),
                record.get('model_used', ''),
                record.get('extraction_confidence', ''),
                record.get('raw_ocr_text', '')
            ]
            writer.writerow(row)
        
        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename=database_records_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Export failed: {str(e)}'
        }), 500

if __name__ == '__main__':
    print("=" * 60)
    print("OCR Service Startup Information:")
    print(f"OCR API: {ocr_status}")
    if ocr_status == "Connected":
        print(f"  - API Address: {OCR_API_BASE}")
        print(f"  - Model Name: {OCR_MODEL}")
    print(f"Backup OCR: {'Tesseract available' if backup_ocr == 'tesseract' else 'Basic processing only'}")
    print(f"Supported formats: {', '.join(ALLOWED_EXTENSIONS)}")
    print(f"Maximum file size: {MAX_FILE_SIZE // (1024*1024)}MB")
    print("=" * 60)
    
    app.run(debug=True, host='0.0.0.0', port=5000)
